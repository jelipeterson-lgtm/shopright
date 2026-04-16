"""
Excel generation — template-copy approach.
Downloads master templates from Dropbox, copies them, writes visit data.
"""
import os
import io
import httpx
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime, date
from collections import defaultdict

from ingest_stores import normalize_dropbox_url

DROPBOX_SHOPFILE_URL = normalize_dropbox_url(os.getenv("DROPBOX_SHOPFILE_TEMPLATE_URL", ""))
DROPBOX_INVOICE_URL = normalize_dropbox_url(os.getenv("DROPBOX_INVOICE_TEMPLATE_URL", ""))

# Cache templates in memory
_template_cache = {}


def _download_template(url, key):
    """Download and cache template. Always return a fresh BytesIO copy."""
    if key not in _template_cache:
        r = httpx.get(url, follow_redirects=True, timeout=30)
        r.raise_for_status()
        _template_cache[key] = r.content
    return io.BytesIO(bytes(_template_cache[key]))


def _format_date(date_str):
    """Convert YYYY-MM-DD to MM/DD/YY."""
    if not date_str:
        return ""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%m/%d/%y")
    except ValueError:
        return date_str


def _format_time(time_str):
    """Convert HH:MM:SS or HH:MM to hh:mm AM/PM."""
    if not time_str:
        return ""
    try:
        # Handle HH:MM:SS or HH:MM
        parts = time_str.split(":")
        hour = int(parts[0])
        minute = int(parts[1])
        period = "AM" if hour < 12 else "PM"
        if hour == 0:
            hour = 12
        elif hour > 12:
            hour -= 12
        return f"{hour}:{minute:02d} {period}"
    except (ValueError, IndexError):
        return time_str


def generate_shop_file(visits, first_name):
    """Generate Shop File .xlsx from a list of Complete visits for one ISO week."""
    template = _download_template(DROPBOX_SHOPFILE_URL, "shopfile")
    wb = load_workbook(template)
    ws = wb.active

    # Clear ALL existing data rows thoroughly (keep row 1 headers only)
    for row in range(2, ws.max_row + 1):
        for col in range(1, 41):
            ws.cell(row, col).value = None

    # Write visit data starting at row 2
    for i, v in enumerate(visits):
        row = i + 2

        # Col 1: Fail Count — count of "Fail" across eval fields (blank if 0)
        fail_count = 0
        eval_fields = [
            "eval_engaging", "eval_greeting", "eval_one_no", "eval_pushy",
            "eval_clogging", "eval_leaning", "eval_food_drink", "eval_dress_code",
            "eval_name_badge", "eval_badge_location_pass", "eval_other_area",
            "eval_other_store_areas", "eval_soft_selling",
        ]
        for f in eval_fields:
            if v.get(f) == "Fail":
                fail_count += 1
        # Only write fail count if > 0, otherwise leave blank
        if fail_count > 0:
            ws.cell(row, 1, fail_count)

        right_align = Alignment(horizontal='right')

        ws.cell(row, 2, v.get("retailer_name") or "")
        ws.cell(row, 3, v.get("program") or "")
        ws.cell(row, 4, v.get("store_number") or "")
        ws.cell(row, 5, v.get("city") or "")
        ws.cell(row, 6, v.get("state") or "")
        date_cell = ws.cell(row, 7, _format_date(v.get("visit_date")))
        date_cell.alignment = right_align
        time_cell = ws.cell(row, 8, _format_time(v.get("visit_time")))
        time_cell.alignment = right_align

        reps = v.get("reps_present") or ""
        ws.cell(row, 9, reps)

        # If Reps Present = Fail, cols 10-39 are blank, only Visit Recap
        if reps == "Fail":
            ws.cell(row, 40, v.get("visit_recap") or "")
            continue

        # If Reps Present = N/A, all eval fields output N/A, only Visit Recap has content
        if reps == "N/A":
            for c in range(10, 40):
                ws.cell(row, c, "N/A")
            ws.cell(row, 40, v.get("visit_recap") or "")
            continue

        ws.cell(row, 10, v.get("rep_names") or "")
        ws.cell(row, 11, v.get("rep_description") or "")

        # Col 12: Less than 4 reps present? — Pass if <= 4, Fail if > 4
        rep_count = v.get("rep_count")
        if rep_count is not None:
            ws.cell(row, 12, "Pass" if rep_count <= 4 else "Fail")

        # Col 13: rep count + reason if > 4
        if rep_count is not None:
            reason = v.get("rep_count_reason") or ""
            if rep_count > 4 and reason:
                count_cell = ws.cell(row, 13, f"{rep_count} - {reason}")
            else:
                count_cell = ws.cell(row, 13, str(rep_count))
            count_cell.alignment = right_align

        ws.cell(row, 14, v.get("eval_engaging") or "Pass")
        ws.cell(row, 15, v.get("eval_engaging_comment") or "")
        ws.cell(row, 16, v.get("eval_greeting") or "Pass")
        ws.cell(row, 17, v.get("eval_greeting_comment") or "")
        ws.cell(row, 18, v.get("eval_one_no") or "Pass")
        ws.cell(row, 19, v.get("eval_one_no_comment") or "")
        ws.cell(row, 20, v.get("eval_pushy") or "Pass")
        ws.cell(row, 21, v.get("eval_pushy_comment") or "")
        ws.cell(row, 22, v.get("eval_clogging") or "Pass")
        ws.cell(row, 23, v.get("eval_clogging_comment") or "")
        ws.cell(row, 24, v.get("eval_leaning") or "Pass")
        ws.cell(row, 25, v.get("eval_leaning_comment") or "")
        ws.cell(row, 26, v.get("eval_food_drink") or "Pass")
        ws.cell(row, 27, v.get("eval_food_drink_comment") or "")
        ws.cell(row, 28, v.get("eval_dress_code") or "Pass")
        ws.cell(row, 29, v.get("eval_dress_code_comment") or "")
        ws.cell(row, 30, v.get("eval_name_badge") or "Pass")
        ws.cell(row, 31, v.get("eval_name_badge_comment") or "")
        ws.cell(row, 32, v.get("eval_badge_location_pass") or "Pass")
        ws.cell(row, 33, v.get("eval_badge_where") or "")
        ws.cell(row, 34, v.get("eval_other_area") or "Pass")
        ws.cell(row, 35, v.get("eval_other_area_comment") or "")
        ws.cell(row, 36, v.get("eval_other_store_areas") or "Pass")
        ws.cell(row, 37, v.get("eval_other_store_areas_comment") or "")
        ws.cell(row, 38, v.get("eval_soft_selling") or "N/A")
        ws.cell(row, 39, v.get("eval_resource_guide") or "N/A")
        ws.cell(row, 40, v.get("visit_recap") or "")

    # Determine filename: last shopping day of the visits
    visit_dates = [v.get("visit_date", "") for v in visits if v.get("visit_date")]
    if visit_dates:
        last_day = max(visit_dates)
        try:
            dt = datetime.strptime(last_day, "%Y-%m-%d")
            date_str = dt.strftime("%m.%d.%y")
        except ValueError:
            date_str = last_day
    else:
        date_str = datetime.now().strftime("%m.%d.%y")

    filename = f"Shop File {first_name} {date_str}.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, filename


def _format_phone(phone):
    """Format phone number to (555) 555-5555."""
    if not phone:
        return ""
    digits = ''.join(c for c in phone if c.isdigit())
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == '1':
        return f"({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return phone


def generate_invoice(visits, mileage_entries, profile, year=None, month=None):
    """
    Generate Invoice .xlsx.
    visits: list of Complete visits for the month
    mileage_entries: list of {date: str, miles: number}
    profile: user profile dict
    """
    from openpyxl.styles import Font, Border, Side

    template = _download_template(DROPBOX_INVOICE_URL, "invoice")
    wb = load_workbook(template)
    ws = wb.active

    data_font = Font(name='Arial', size=10)
    thin_border = Border(
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    no_border = Border()

    full_name = profile.get("full_name", "")
    phone = _format_phone(profile.get("phone", ""))
    email = profile.get("report_email", "")

    # Row 1: Invoice number as YYMM
    if year and month:
        invoice_id = f"{year % 100:02d}{month:02d}"
    else:
        now = datetime.now()
        invoice_id = f"{now.year % 100:02d}{now.month:02d}"
    ws.cell(1, 2, f"INVOICE #{invoice_id}")

    # Rows 2-4: User info (keep original template fonts)
    ws.cell(2, 2, full_name)
    ws.cell(3, 2, profile.get("home_address", ""))
    ws.cell(4, 2, f"{phone}  {email}")

    # Row 6: Date (mm/dd/yyyy format)
    ws.cell(6, 3, datetime.now().strftime("%m/%d/%Y"))

    # Row 82: "Make all checks payable to"
    ws.cell(82, 2, f"Make all checks payable to {full_name}")

    # Row 84: Contact info with formatted phone
    ws.cell(84, 2, f"{full_name}, {phone}, {email}")

    # Clear ALL existing mileage and vendor rows — values AND borders
    for row in range(14, 77):
        for col in range(1, 11):
            cell = ws.cell(row, col)
            cell.value = None
            cell.border = no_border
            cell.font = data_font

    current_row = 14

    # Write single mileage line — total miles for the month
    mileage_rate = profile.get("mileage_rate", 0.70)
    total_miles = sum(e.get("miles", 0) for e in mileage_entries if e.get("miles", 0) > 0)
    if total_miles > 0:
        for col in range(2, 11):
            ws.cell(current_row, col).border = thin_border
        ws.cell(current_row, 5, "Mileage").font = data_font
        ws.cell(current_row, 8, mileage_rate).font = data_font
        ws.cell(current_row, 9, total_miles).font = data_font
        ws.cell(current_row, 10, f"=I{current_row}*H{current_row}").font = data_font
        current_row += 1

    # Calculate vendor pricing: $50 first vendor per stop per day, $15 additional
    stops = defaultdict(list)
    for v in visits:
        key = (v.get("visit_date", ""), v.get("store_number", ""), v.get("retailer_name", ""))
        stops[key].append(v)

    sorted_stops = sorted(stops.items(), key=lambda x: (x[0][0], x[0][2], x[0][1]))

    for (visit_date, store_num, retailer), stop_visits in sorted_stops:
        for j, v in enumerate(stop_visits):
            price = 50 if j == 0 else 15
            # Apply consistent thin borders across all columns
            for col in range(2, 11):
                ws.cell(current_row, col).border = thin_border
            ws.cell(current_row, 2, store_num).font = data_font
            ws.cell(current_row, 3, v.get("city", "")).font = data_font
            ws.cell(current_row, 4, v.get("retailer_name", "")).font = data_font
            ws.cell(current_row, 5, v.get("program", "")).font = data_font
            ws.cell(current_row, 6, _format_date(visit_date)).font = data_font
            ws.cell(current_row, 10, price).font = data_font
            current_row += 1

    # Update subtotal formula
    ws.cell(77, 10, f"=SUM(J14:J{current_row - 1})")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output, invoice_id
