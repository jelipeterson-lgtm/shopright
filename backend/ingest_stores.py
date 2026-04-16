"""
Download Book1.xlsx from Dropbox, geocode addresses, and load into stores table.
Also loads program list into a programs table.

- Can be run manually: python3 ingest_stores.py
- Can be called on backend startup: check_and_ingest() HEADs the Dropbox URL,
  compares Last-Modified against a cached value, and only re-ingests if the
  file has changed.
"""
import os
import time
import httpx
from openpyxl import load_workbook
from dotenv import load_dotenv
from db import supabase_admin

load_dotenv()


def normalize_dropbox_url(url):
    """Ensure a Dropbox shared URL uses dl=1 for direct download.

    Handles three cases:
    - `...dl=0`  -> `...dl=1`
    - `...dl=1`  -> unchanged
    - no dl param -> append `dl=1`
    """
    if not url:
        return url
    if "dl=0" in url:
        return url.replace("dl=0", "dl=1")
    if "dl=1" in url:
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}dl=1"


DROPBOX_URL = normalize_dropbox_url(os.getenv("DROPBOX_STORES_URL", ""))

# Cache file tracks the Dropbox Last-Modified we last ingested, so we can skip
# re-downloading when the file hasn't changed. /tmp is wiped on Render cold
# starts, so cold starts will re-ingest — the geocoding step below skips
# already-geocoded stores, which keeps re-ingests fast.
LAST_MODIFIED_CACHE = "/tmp/book1_last_modified.txt"
BOOK1_PATH = "/tmp/Book1.xlsx"


def _read_cached_last_modified():
    try:
        with open(LAST_MODIFIED_CACHE) as f:
            return f.read().strip() or None
    except (FileNotFoundError, OSError):
        return None


def _write_cached_last_modified(value):
    if not value:
        return
    try:
        with open(LAST_MODIFIED_CACHE, "w") as f:
            f.write(value)
    except OSError as e:
        print(f"Failed to write Last-Modified cache: {e}")


def _fetch_remote_last_modified():
    """HEAD the Dropbox URL and return its Last-Modified header, or None."""
    if not DROPBOX_URL:
        return None
    try:
        r = httpx.head(DROPBOX_URL, follow_redirects=True, timeout=15)
        if r.status_code >= 400:
            print(f"HEAD {DROPBOX_URL} returned {r.status_code}")
            return None
        return r.headers.get("last-modified")
    except Exception as e:
        print(f"HEAD request for Dropbox stores URL failed: {e}")
        return None


def download_book1():
    print("Downloading Book1.xlsx from Dropbox...")
    r = httpx.get(DROPBOX_URL, follow_redirects=True, timeout=60)
    r.raise_for_status()
    with open(BOOK1_PATH, "wb") as f:
        f.write(r.content)
    print(f"Downloaded {len(r.content)} bytes to {BOOK1_PATH}")
    return BOOK1_PATH


def geocode_address(address, city, state, zip_code):
    full_address = f"{address}, {city}, {state} {zip_code}"
    try:
        r = httpx.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": full_address, "format": "json", "limit": 1},
            headers={"User-Agent": "ShopRight/1.0"},
            timeout=10,
        )
        data = r.json()
        if data:
            return float(data[0]["lat"]), float(data[0]["lon"])
    except Exception as e:
        print(f"  Geocode failed for {full_address}: {e}")
    return None, None


def _load_existing_geocodes():
    """Return {(address, city, state, zip_code): (lat, lon)} for rows already geocoded."""
    existing = {}
    try:
        resp = (
            supabase_admin.table("stores")
            .select("address,city,state,zip_code,latitude,longitude")
            .execute()
        )
        for row in (resp.data or []):
            if row.get("latitude") and row.get("longitude"):
                key = (
                    (row.get("address") or "").strip(),
                    (row.get("city") or "").strip(),
                    (row.get("state") or "").strip(),
                    str(row.get("zip_code") or "").strip(),
                )
                existing[key] = (row["latitude"], row["longitude"])
    except Exception as e:
        print(f"Failed to load existing geocodes: {e}")
    return existing


def parse_and_load(path):
    wb = load_workbook(path)

    # --- Load Programs ---
    if "Program" in wb.sheetnames:
        ws_prog = wb["Program"]
        programs = []
        for row in range(2, ws_prog.max_row + 1):
            code = ws_prog.cell(row, 1).value
            if code and code.strip():
                programs.append(code.strip())
        print(f"Found {len(programs)} programs")

        # Store programs in database
        for prog in programs:
            supabase_admin.table("programs").upsert(
                {"code": prog}, on_conflict="code"
            ).execute()
        print(f"Programs loaded: {programs}")
    else:
        print("No Program worksheet found")

    # --- Load Stores from Retail worksheet ---
    ws_name = "Retail" if "Retail" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    seen = set()
    stores = []

    for row in range(2, ws.max_row + 1):
        retailer = ws.cell(row, 1).value
        store_num = str(ws.cell(row, 2).value or "").strip()

        if not retailer or not store_num:
            continue

        key = (retailer.strip(), store_num)
        if key in seen:
            continue
        seen.add(key)

        stores.append({
            "retailer_name": retailer.strip(),
            "store_number": store_num,
            "address": (ws.cell(row, 3).value or "").strip(),
            "city": (ws.cell(row, 4).value or "").strip(),
            "state": (ws.cell(row, 5).value or "").strip(),
            "zip_code": str(ws.cell(row, 6).value or "").strip(),
        })

    print(f"Parsed {len(stores)} unique stores (from {ws.max_row - 1} rows)")

    # Pre-populate geocode cache from DB so we only call Nominatim for new
    # addresses. Dramatically speeds up re-ingests after cold starts.
    geocoded = _load_existing_geocodes()
    print(f"Loaded {len(geocoded)} existing geocodes from database")

    # Geocode each unique address not already known
    for store in stores:
        addr_key = (store["address"], store["city"], store["state"], store["zip_code"])
        if addr_key not in geocoded:
            lat, lon = geocode_address(*addr_key)
            geocoded[addr_key] = (lat, lon)
            if lat:
                print(f"  Geocoded: {store['retailer_name']} #{store['store_number']} -> {lat}, {lon}")
            else:
                print(f"  FAILED: {store['retailer_name']} #{store['store_number']}")
            time.sleep(1.1)  # Nominatim rate limit

        store["latitude"], store["longitude"] = geocoded[addr_key]

    # Upsert into database
    print("Loading into database...")
    for store in stores:
        supabase_admin.table("stores").upsert(
            store, on_conflict="retailer_name,store_number"
        ).execute()

    print(f"Done. {len(stores)} stores loaded.")


def check_and_ingest(force=False):
    """Check Dropbox Last-Modified; re-ingest Book1.xlsx only if changed.

    Safe to call on every backend startup. Swallows all errors so a Dropbox
    outage can't take the API down.
    """
    if not DROPBOX_URL:
        print("DROPBOX_STORES_URL not set — skipping store directory check")
        return

    try:
        remote = _fetch_remote_last_modified()
        cached = _read_cached_last_modified()

        if not force and remote and cached and remote == cached:
            print(f"Store directory unchanged (Last-Modified: {remote}) — skipping re-ingest")
            return

        print(
            f"Store directory changed — re-ingesting "
            f"(remote Last-Modified: {remote}, cached: {cached})"
        )
        path = download_book1()
        parse_and_load(path)
        _write_cached_last_modified(remote)
    except Exception as e:
        print(f"check_and_ingest failed: {e}")


if __name__ == "__main__":
    # Manual runs always force re-ingest regardless of cached Last-Modified.
    check_and_ingest(force=True)
